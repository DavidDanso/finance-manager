from django.http import HttpResponse
from accounts.models import Profile
from .forms import UserCreationForm
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from reports.forms import ReportCreationForm, ReportEditForm, ReviewReportForm
from django.http import HttpResponse
from openpyxl import Workbook
from django.db.models import Sum
from reports.models import Report
from openpyxl.utils import get_column_letter
from io import BytesIO
from django.contrib import messages
from reports.filters import ReportFilter
from .forms import ProfileUpdateForm, UserUpdateForm


# admin dashboard view
@login_required
def admin_dashboard(request):
    if request.user.role != 'admin':
        return redirect('login')

    user = request.user.profile
    form = ReportCreationForm()

    # Apply the filter
    reports = Report.objects.select_related('account_owner').all()
    report_filter = ReportFilter(request.GET, queryset=reports)
    reports = report_filter.qs.select_related('account_owner')

    # Get total pending reports count
    pending_reports_count = reports.filter(status='pending').count()

    # Get total amount of approved reports
    total_amount = reports.filter(status='approve').aggregate(Sum('payment'))['payment__sum'] or 0

    # Fetch reports excluding those with 'pending' status
    other_reports = reports.exclude(status='pending')

    # Fetch first 5 pending reports
    pending_reports = reports.filter(status='pending')[:5]

    if request.method == "POST":
        form = ReportCreationForm(request.POST)
        if form.is_valid():
            report = form.save(commit=False)
            report.account_owner = user
            if request.user.role == "admin":
                report.status = 'approve'
            report.save()
            messages.success(request, '📄 Report Created Successfully')
            return redirect('dashboard')

    context = {
        'form': form,
        'reports': other_reports,
        'filter': report_filter,
        'reports_count': pending_reports_count,
        'total_amount': total_amount,
        'pending_reports': pending_reports,
    }

    if request.htmx:
        return render(request, 'partials/base-reports-list.html', context)

    return render(request, 'u_admin/admin_dashboard.html', context)






# review report view
def review_report(request, report_id):
    report = Report.objects.get(id=report_id)
    form = ReviewReportForm(instance=report)

    if request.method == 'POST':
        form = ReviewReportForm(request.POST, instance=report)
        if form.is_valid():
            messages.success(request, 'Report Review Completed Successfully!')
            form.save()
            return redirect('pending-reports')

    context = {'form': form, 'report': report}
    return render(request, 'u_admin/review-report.html', context)




# edit_report view
@login_required
def edit_report(request, pk):
    # Get the report object based on the provided primary key
    report = Report.objects.get(id=pk)
    
    # Create a form instance pre-populated with the report data
    form = ReportEditForm(instance=report)

    # Handle form submission
    if request.method == "POST":
        # Check if the user wants to delete the report
        if 'delete_report' in request.POST:
            # Delete the report and redirect to the dashboard
            report.delete()
            messages.success(request, 'Report deleted successfully')
            return redirect('dashboard')
        
        # Create a form instance with the submitted data and the report instance
        form = ReportEditForm(request.POST, instance=report)
        
        # Check if the form is valid
        if form.is_valid():
            # Save the updated report and display a success message
            messages.success(request, 'Report successfully updated! ✅')
            report.save()
            return redirect('dashboard')
        
        else:
            # Display an error message if the form is invalid
            messages.error(request, 'Report update failed. 🚫')
        
    # Prepare the context for rendering the template
    context = {'form': form, 'report': report}
    
    # Render the edit-reports.html template with the context
    return render(request, 'edit-reports.html', context)




# pending-reports view
def pending_reports(request):
    # Fetch all reports with the status 'pending'
    pending_reports = Report.objects.filter(status='pending')

    # Count the number of pending reports
    pending_count = pending_reports.count()

    # Prepare the context for rendering the template
    context = {
        'pending_reports': pending_reports,  # List of pending reports
        'count': pending_count,      # Total count of pending reports
    }
    # Render the 'pending-reports.html' template with the context
    return render(request, 'u_admin/pending-reports.html', context)




# user_management view
@login_required
def user_management(request):
    if request.user.role != 'admin':
        return redirect('login')
    
    users = Profile.objects.all()
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.username = user.username.lower()
            user.save()
            messages.success(request, 'User created successful')
            return redirect('users')
    else:
        form = UserCreationForm()

    # Fetch all reports with the status 'pending'[ notification ]
    pending_reports = Report.objects.filter(status='pending')

    context = {'users': users, 'form': form, 'pending_reports': pending_reports}    
    return render(request, 'u_admin/user_management.html', context)




# download_report view
def download_report(request, report_id):
    report = get_object_or_404(Report, id=report_id)

    # Create a workbook and add a worksheet
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Report"

    # Define the headers
    headers = ["Username", "Date", "Description", "Main Category", "Sub Category", "Payment"]
    worksheet.append(headers)

    # Add the report data
    formatted_date = report.created_time_stamp.strftime('%d %B, %Y') if report.created_time_stamp else None
    report_data = [
        report.account_owner.username,
        formatted_date,
        report.description,
        report.main_category,
        report.sub_category,
        str(report.payment)
    ]
    worksheet.append(report_data)

    # Adjust column widths
    column_widths = []
    for row in worksheet.iter_rows(values_only=True):
        for i, cell in enumerate(row):
            if len(column_widths) > i:
                if len(str(cell)) > column_widths[i]:
                    column_widths[i] = len(str(cell))
            else:
                column_widths += [len(str(cell))]

    for i, column_width in enumerate(column_widths, 1):
        worksheet.column_dimensions[get_column_letter(i)].width = column_width + 2

    # Save the workbook to a BytesIO stream
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    # Set the response as an Excel file
    response = HttpResponse(
        content=output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=report_{report_id}.xlsx'

    return response




# update user view
@login_required
def update_user(request, pk):
    # Check if the current user is an admin
    if request.user.role != 'admin':
        return redirect('login')

    # Get the user profile, or return 404 if not found
    profile = get_object_or_404(Profile, id=pk)
    user = profile.user

    if request.method == 'POST':
        # Check if the request is for deleting the user
        if 'delete_user' in request.POST:
            profile.delete()
            messages.success(request, 'User deleted successful')
            return redirect('users')

        # Process the user update forms
        user_form = UserUpdateForm(request.POST, instance=user)
        profile_form = ProfileUpdateForm(request.POST, instance=profile)
        if user_form.is_valid() and profile_form.is_valid():
            user_form.save()
            profile_form.save()
            messages.success(request, 'User updated successfully')
            return redirect('users')
    else:
        # Initialize forms with existing user and profile data
        user_form = UserUpdateForm(instance=user)
        profile_form = ProfileUpdateForm(instance=profile)

    # Context to be passed to the template
    context = {
        'user_form': user_form,
        'profile_form': profile_form,
        'user': user,
        'profile': profile
    }

    # Render the edit user template
    return render(request, 'u_admin/edit-user.html', context)

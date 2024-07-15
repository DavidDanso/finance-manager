from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from reports.forms import ReportCreationForm, ReportEditForm
from django.http import HttpResponse
from openpyxl import Workbook
from reports.models import Report
from django.contrib import messages
from openpyxl.utils import get_column_letter
from io import BytesIO
from reports.filters import ReportFilter
import pandas as pd
from reports.forms import UploadFileForm
from accounts.models import Profile
from openpyxl import load_workbook
from django.db.models import Count, Q
from datetime import datetime
from django.core.paginator import Paginator
from django.conf import settings

# dashboard view
@login_required(login_url='login')
def accountant_dashboard(request):
    user = request.user.profile

    # Fetch all reports for the user in a single query using select_related
    all_reports = Report.objects.select_related('account_owner').filter(account_owner=user)

    # Exclude reports with status 'correction' from the already fetched reports
    other_reports = [report for report in all_reports if report.status != 'correction'][:5]

    # Get the latest correction report from the already fetched reports
    new_correction = next((report for report in all_reports if report.status == 'correction'), None)

    # Aggregate counts in a single query
    status_counts = all_reports.aggregate(
        approve_count=Count('id', filter=Q(status='approve')),
        reject_count=Count('id', filter=Q(status='reject')),
        pending_count=Count('id', filter=Q(status='pending')),
        correction_count=Count('id', filter=Q(status='correction')),
    )

    # Extract the counts from the aggregated result
    approve_count = status_counts['approve_count']
    reject_count = status_counts['reject_count']
    pending_count = status_counts['pending_count']
    correction_count = status_counts['correction_count']

    # Fetch reports with the status 'correction, approve, reject' for the logged-in user [notification]
    notification_reports = [report for report in all_reports if report.status != 'pending'][:5]


    # get current date
    current_date = datetime.now().strftime("%B %d, %Y")

    context = {
        'reports': all_reports,
        'approve_count': approve_count,
        'reject_count': reject_count,
        'pending_count': pending_count,
        'correction_count': correction_count,
        'current_date': current_date,
        'correction': new_correction,
        'notification_reports': notification_reports,
        'other_reports': other_reports,
    }
    
    return render(request, 'accountant/a_dashboard.html', context)





# corrections view
@login_required(login_url='login')
def corrections_page(request):
    # Get the profile of the logged-in user
    user = request.user.profile

    # Fetch reports with the status 'correction' and other relevant statuses for the logged-in user in a single query
    all_reports = Report.objects.filter(account_owner=user).select_related('account_owner')

    # Filter reports with status 'correction'
    correction_reports = [report for report in all_reports if report.status == 'correction']

    # Count the number of correction reports
    correction_count = len(correction_reports)

    # Fetch reports with the status 'correction', 'approve', or 'reject' for notifications
    notification_reports = [report for report in all_reports if report.status != 'pending'][:5]

    # Prepare the context for rendering the template
    context = {
        'correction_reports': correction_reports,  # List of correction reports
        'correction_count': correction_count,      # Total count of correction reports
        'notification_reports': notification_reports
    }

    # Render the 'corrections.html' template with the context
    return render(request, 'accountant/corrections.html', context)





# Ensure the request is a POST request
def make_correction(request, report_id):
    user = request.user.profile
    report = user.report_set.get(id=report_id)
    form = ReportEditForm(instance=report)

    if request.method == "POST":
        form = ReportEditForm(request.POST, instance=report)
        report = form.save(commit=False)
        report.status = 'pending'
        messages.success(request, 'Report successfully sent to Admin!')
        report.save()
        return redirect('corrections')
    
    context = {'form': form, 'report': report}
    return render(request, 'accountant/report-correction.html', context)




# upload_file view
@login_required(login_url='login')
def upload_file(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            if not file.name.endswith('.xlsx'):
                messages.error(request, 'The uploaded file is not an Excel file.')
                return redirect('upload_file')
            try:
                # Verify the file is an Excel file by attempting to load it
                load_workbook(file)
                data = pd.read_excel(file, engine='openpyxl')
                for index, row in data.iterrows():
                    try:
                        account_owner = Profile.objects.get(username=row['username'])
                    except Profile.DoesNotExist:
                        messages.error(request, f"Profile with username {row['username']} does not exist.")
                        continue
                    Report.objects.create(
                        account_owner=account_owner,
                        description=row['description'],
                        main_category=row['main category'],
                        sub_category=row['sub category'],
                        status=row['status'],
                        payment=row['payment'],
                    )
                messages.success(request, 'File uploaded and reports created successfully.')
                return redirect('upload_file')
            except Exception as e:
                messages.error(request, f'Error processing file: {e}')
    else:
        form = UploadFileForm()
    return render(request, 'accountant/upload-reports.html', {'form': form})




# create_report view
@login_required(login_url='login')
def create_report(request):
    if request.user.role != 'accountant':
        return redirect('login')
    
    user = request.user.profile
    form = ReportCreationForm()

    # Fetch all Report objects from the database with their related account_owner in a single query
    all_reports = Report.objects.select_related('account_owner').all()
    
    # Apply the filter
    report_filter = ReportFilter(request.GET, queryset=all_reports)
    reports = report_filter.qs

    # Paginator - get the correct page number from the request
    paginator = Paginator(reports, settings.PAGE_SIZE)
    page_number = request.GET.get('page', 1)
    reports_page = paginator.get_page(page_number)

    # Fetch reports with the relevant statuses for the logged-in user in a single query
    logged_user_reports = Report.objects.filter(account_owner=user).select_related('account_owner')
    
    # Fetch reports with status 'correction', 'approve', or 'reject' for notifications
    notification_reports = logged_user_reports.exclude(status='pending')[:5]
    
    # Create reports
    if request.method == "POST":
        form = ReportCreationForm(request.POST)
        if form.is_valid():
            report = form.save(commit=False)
            report.account_owner = user
            report.save()
            messages.success(request, 'Report successfully created! ✅')
            return redirect('reports')
        
    context = {
        'form': form,
        'filter': report_filter,
        'list': ['reject', 'approve', 'pending'],
        'user': user,
        'notification_reports': notification_reports,
        'reports': reports_page,
    }

    if request.htmx:
        return render(request, 'partials/reports-list.html', context)

    return render(request, 'accountant/reports.html', context)





# edit_report view
@login_required(login_url='login')
def edit_report(request, pk):
    user = request.user.profile
    report = user.report_set.get(id=pk)
    form = ReportEditForm(instance=report)

    #
    if request.method == "POST":
        form = ReportEditForm(request.POST, instance=report)
        if form.is_valid():
            messages.success(request, 'Report successfully updated! ✅')
            report.save()
        else:
            messages.error(request, 'Report update failed. 🚫')
        
    context = {'form': form, 'report': report}
    return render(request, 'edit-reports.html', context)




# download_report view
def download_report(request, report_id):
    report = get_object_or_404(Report, id=report_id)

    # Create a workbook and add a worksheet
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Report"

    # Define the headers
    headers = ["username", "date", "description", "main category", "sub category", "payment"]
    worksheet.append(headers)

    # Add the report data
    formatted_date = report.created_time_stamp.strftime('%d %B, %Y') if report.created_time_stamp else None
    report_data = [
        report.account_owner.username,
        formatted_date,
        report.description,
        report.main_category,
        report.sub_category,
        str(report.payment)
    ]
    worksheet.append(report_data)

    # Adjust column widths
    column_widths = []
    for row in worksheet.iter_rows(values_only=True):
        for i, cell in enumerate(row):
            if len(column_widths) > i:
                if len(str(cell)) > column_widths[i]:
                    column_widths[i] = len(str(cell))
            else:
                column_widths += [len(str(cell))]

    for i, column_width in enumerate(column_widths, 1):
        worksheet.column_dimensions[get_column_letter(i)].width = column_width + 2

    # Save the workbook to a BytesIO stream
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    # Set the response as an Excel file
    response = HttpResponse(
        content=output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=report_{report_id}.xlsx'

    return response

